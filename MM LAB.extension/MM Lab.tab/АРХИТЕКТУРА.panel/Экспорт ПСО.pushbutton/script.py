#! python3
# -*- coding: utf-8 -*-
"""Экспорт ПСО в Excel

Экспортирует спецификацию помещений (ПСО) из Revit в Excel
с объединением ячеек по квартирам.

Совместимость: Revit 2020 / 2022 / 2024 (CPython3 PyRevit)
Зависимости: openpyxl (устанавливается в MMLab_TOOLS/lib)
"""

__title__ = "Экспорт\nПСО"
__author__ = "GENPRO LAB"
__doc__ = (
    "Экспорт спецификации помещений (ПСО) в Excel "
    "с группировкой и объединением ячеек по квартирам."
)

# ==============================================================================
# IMPORTS
# ==============================================================================
import clr
import sys
import os
from collections import OrderedDict
from datetime import datetime

clr.AddReference('RevitAPI')
clr.AddReference('RevitAPIUI')
import Autodesk.Revit.DB as DB
import Autodesk.Revit.UI as UI

# WinForms для диалога выбора папки (pyrevit.forms не работает под CPython3)
clr.AddReference('System.Windows.Forms')
from System.Windows.Forms import (
    SaveFileDialog,
    DialogResult,
)

# ==============================================================================
# UI helpers — замена pyrevit.forms для CPython3
# ==============================================================================

def alert(message, title="Экспорт ПСО"):
    """Показывает диалог через Revit TaskDialog."""
    td = UI.TaskDialog(title)
    td.MainContent = message
    td.CommonButtons = UI.TaskDialogCommonButtons.Ok
    td.Show()


def save_file(title="Сохранить файл", default_name="", file_filter=""):
    """Диалог сохранения файла через WinForms SaveFileDialog."""
    dlg = SaveFileDialog()
    dlg.Title = title
    dlg.FileName = default_name
    dlg.Filter = file_filter
    dlg.OverwritePrompt = True
    if dlg.ShowDialog() == DialogResult.OK:
        return dlg.FileName
    return None


# ==============================================================================
# VENDOR path (openpyxl)
# ==============================================================================
SCRIPT_DIR     = os.path.dirname(__file__)
EXTENSION_ROOT = os.path.normpath(os.path.join(SCRIPT_DIR, "..", "..", "..", ".."))
VENDOR_DIR     = os.path.join(EXTENSION_ROOT, "lib")

if os.path.isdir(VENDOR_DIR) and VENDOR_DIR not in sys.path:
    sys.path.insert(0, VENDOR_DIR)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    alert(
        "Библиотека openpyxl не найдена.\n\n"
        "Установите командой:\n"
        "  pip install openpyxl --target \"{}\"\n\n"
        "и перезагрузите PyRevit.".format(VENDOR_DIR),
        title="Ошибка импорта",
    )
    sys.exit(0)

# ==============================================================================
# Revit context
# ==============================================================================
doc = __revit__.ActiveUIDocument.Document  # noqa: F821

# ==============================================================================
# CONFIGURATION — имена параметров
# ==============================================================================
PARAM_KORPUS            = "GP_23_НомерКорпуса"
PARAM_KV_NUMBER         = "GP_23_НомерКв"
PARAM_NAZNACHENIE       = "GP_23_Назначение"
PARAM_ETAZH             = "GP_01_Этаж_Номер"
PARAM_SECTION           = "GP_23_НомерСекции"
PARAM_PL_KV_BEZ_KOEF   = "GP_23_ПлКвОбщая+НеотБезКоэф_ПСО"
PARAM_PL_KV_S_KOEF      = "GP_23_ПлКвОбщая+Неот_ПСО"
PARAM_KOL_KOMNAT        = "GP_23_КолвоКомнат"
PARAM_PL_ZHILAYA        = "GP_23_ПлКвЖилая_ПСО"
PARAM_VYSOTA            = "Полная высота"

PARAM_ROOM_NUM_IN_KV    = "GP_23_НомерПомКв"
PARAM_ROOM_NAME         = "Имя"
PARAM_PL_ROOM_BEZ_KOEF  = "GP_23_Площадь_ПСО"
PARAM_PL_ROOM_S_KOEF    = "GP_23_ПлощадьСКоэф_ПСО"


# ==============================================================================
# HELPERS — параметры
# ==============================================================================

def get_param_as_string(room, param_name):
    """Безопасное чтение значения параметра как строки."""
    p = None
    if param_name == "Имя":
        p = room.get_Parameter(DB.BuiltInParameter.ROOM_NAME)
    else:
        p = room.LookupParameter(param_name)

    if not p or not p.HasValue:
        return ""

    storage = p.StorageType
    if storage == DB.StorageType.String:
        return p.AsString() or ""
    elif storage == DB.StorageType.Double:
        vs = p.AsValueString()
        return vs if vs else str(round(p.AsDouble(), 4))
    elif storage == DB.StorageType.Integer:
        vs = p.AsValueString()
        return vs if vs else str(p.AsInteger())
    elif storage == DB.StorageType.ElementId:
        eid = p.AsElementId()
        if not eid:
            return ""
        try:
            return str(eid.Value)        # Revit 2024+
        except AttributeError:
            return str(eid.IntegerValue)  # Revit 2020/2022
    return ""


def try_parse_number(text):
    """Пробует привести строковое значение к числу для Excel."""
    if not text or not isinstance(text, str):
        return text
    cleaned = text.strip().replace(",", ".").replace("\xa0", "")
    for suffix in [" м²", " м2", " мм", " м", " mm", " m²", " m"]:
        if cleaned.lower().endswith(suffix):
            cleaned = cleaned[: -len(suffix)].strip()
            break
    try:
        val = float(cleaned)
        if val == int(val) and "." not in cleaned:
            return int(val)
        return val
    except (ValueError, TypeError):
        return text


# ==============================================================================
# DATA COLLECTION
# ==============================================================================

def collect_placed_rooms():
    """Возвращает список размещённых Room-элементов (Area > 0)."""
    elements = (
        DB.FilteredElementCollector(doc)
        .OfCategory(DB.BuiltInCategory.OST_Rooms)
        .WhereElementIsNotElementType()
        .ToElements()
    )
    result = []
    for r in elements:
        if r and r.Area > 0:
            result.append(r)
    return result


def build_apartment_data(rooms):
    """Группирует помещения по квартирам."""
    apartments = {}
    for room in rooms:
        korpus  = get_param_as_string(room, PARAM_KORPUS)
        kv_num  = get_param_as_string(room, PARAM_KV_NUMBER)
        etazh   = get_param_as_string(room, PARAM_ETAZH)
        section = get_param_as_string(room, PARAM_SECTION)
        key = (korpus, kv_num, etazh, section)

        if key not in apartments:
            apartments[key] = {
                "korpus":         korpus,
                "kv_number":      kv_num,
                "naznachenie":    get_param_as_string(room, PARAM_NAZNACHENIE),
                "etazh":          etazh,
                "section":        section,
                "pl_kv_bez_koef": get_param_as_string(room, PARAM_PL_KV_BEZ_KOEF),
                "pl_kv_s_koef":   get_param_as_string(room, PARAM_PL_KV_S_KOEF),
                "kol_komnat":     get_param_as_string(room, PARAM_KOL_KOMNAT),
                "pl_zhilaya":     get_param_as_string(room, PARAM_PL_ZHILAYA),
                "vysota":         get_param_as_string(room, PARAM_VYSOTA),
                "rooms": [],
            }

        apartments[key]["rooms"].append({
            "num_in_kv":   get_param_as_string(room, PARAM_ROOM_NUM_IN_KV),
            "name":        get_param_as_string(room, PARAM_ROOM_NAME),
            "pl_bez_koef": get_param_as_string(room, PARAM_PL_ROOM_BEZ_KOEF),
            "pl_s_koef":   get_param_as_string(room, PARAM_PL_ROOM_S_KOEF),
        })

    for apt in apartments.values():
        apt["rooms"].sort(key=lambda r: r["num_in_kv"])

    def _sort_key(item):
        parts = []
        for v in item[0]:
            try:
                parts.append((0, int(v)))
            except (ValueError, TypeError):
                parts.append((1, v or ""))
        return parts

    return OrderedDict(sorted(apartments.items(), key=_sort_key))


# ==============================================================================
# EXCEL
# ==============================================================================

HEADER_TITLES = [
    "Корпус", "Условный\nномер", "Назначение", "Этаж\nрасположения",
    "Номер\nподъезда",
    ("Общая площадь квартиры\n(включая площадь\nбалконов, лоджий, террас,\n"
     "веранд без понижающего\nкоэффициента), м2"),
    ("Общая площадь квартиры\n(включая площадь\nбалконов, лоджий, террас,\n"
     "веранд с понижающими\nкоэффициентами), м2"),
    "Кол-во\nкомнат", "Общая\nжилая\nплощадь, м2", "Высота\nпотолков",
    "Условный номер\nкомнат помещений\nв составе\nквартиры",
    "Комнаты\nв составе квартиры",
    ("Проектная площадь\nкомнат и помещений\n(включая площадь лоджий,\n"
     "балконов, террас и\nверанд без понижающего\nкоэффициента) в составе\nквартиры, кв.м"),
    ("Проектная площадь\nкомнат и помещений\n(включая площадь лоджий,\n"
     "балконов, террас и\nверанд с понижающим\nкоэффициентом) в составе\nквартиры, кв.м"),
]
COL_WIDTHS   = [8, 10, 13, 11, 9, 24, 24, 9, 13, 11, 15, 20, 24, 24]
KV_COL_RANGE = range(1, 11)


def _make_styles():
    thin = Side(style="thin")
    return {
        "border":    Border(left=thin, right=thin, top=thin, bottom=thin),
        "hdr_font":  Font(name="Arial", size=9, bold=True, color="FFFFFF"),
        "hdr_fill":  PatternFill("solid", fgColor="4472C4"),
        "hdr_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "d_font":    Font(name="Arial", size=9),
        "d_center":  Alignment(horizontal="center", vertical="center", wrap_text=True),
        "d_left":    Alignment(horizontal="left",   vertical="center", wrap_text=True),
    }


def _write_header(ws, st, row=1):
    for ci, title in enumerate(HEADER_TITLES, 1):
        c = ws.cell(row=row, column=ci, value=title)
        c.font = st["hdr_font"]; c.fill = st["hdr_fill"]
        c.alignment = st["hdr_align"]; c.border = st["border"]
    ws.row_dimensions[row].height = 95
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _write_apartment(ws, st, row, apt):
    room_list = apt["rooms"]
    n = len(room_list)
    if n == 0:
        return row
    end = row + n - 1
    kv_vals = [
        apt["korpus"], apt["kv_number"], apt["naznachenie"], apt["etazh"],
        apt["section"], apt["pl_kv_bez_koef"], apt["pl_kv_s_koef"],
        apt["kol_komnat"], apt["pl_zhilaya"], apt["vysota"],
    ]
    for ci, val in enumerate(kv_vals, 1):
        cell = ws.cell(row=row, column=ci, value=try_parse_number(val))
        cell.font = st["d_font"]; cell.alignment = st["d_center"]; cell.border = st["border"]
    if n > 1:
        for ci in KV_COL_RANGE:
            ws.merge_cells(start_row=row, start_column=ci, end_row=end, end_column=ci)
    for ri, rm in enumerate(room_list):
        cr = row + ri
        for offset, val in enumerate([rm["num_in_kv"], rm["name"], rm["pl_bez_koef"], rm["pl_s_koef"]]):
            ci = 11 + offset
            cell = ws.cell(row=cr, column=ci, value=try_parse_number(val))
            cell.font = st["d_font"]; cell.border = st["border"]
            cell.alignment = st["d_left"] if ci == 12 else st["d_center"]
    for r in range(row, end + 1):
        for c in range(1, 15):
            ws.cell(row=r, column=c).border = st["border"]
    return end + 1


def export_to_excel(apartments):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ПСО"
    st = _make_styles()
    _write_header(ws, st, row=1)
    cur, count = 2, 0
    for apt in apartments.values():
        cur = _write_apartment(ws, st, cur, apt)
        count += 1
    ws.freeze_panes = "A2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    try:
        ws.sheet_properties.pageSetUpPr = (
            openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        )
    except Exception:
        pass
    return wb, count


# ==============================================================================
# MAIN
# ==============================================================================

try:
    # --- Сбор помещений ---
    rooms = collect_placed_rooms()
    if not rooms:
        alert(
            "В модели не найдено размещённых помещений (Rooms).",
            title="Нет данных",
        )
        sys.exit(0)

    apartments = build_apartment_data(rooms)
    if not apartments:
        alert(
            "Не удалось сгруппировать помещения по квартирам.\n"
            "Проверьте параметр «{}».".format(PARAM_KV_NUMBER),
            title="Нет данных",
        )
        sys.exit(0)

    # --- Выбор файла для сохранения ---
    ts           = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_name = "PSO_Export_{}.xlsx".format(ts)
    save_path = save_file(
        title="Сохранить ПСО",
        default_name=default_name,
        file_filter="Excel (*.xlsx)|*.xlsx",
    )
    if not save_path:
        sys.exit(0)

    # --- Экспорт ---
    wb, total_apt   = export_to_excel(apartments)
    wb.save(save_path)

    total_rooms = sum(len(a["rooms"]) for a in apartments.values())
    alert(
        "Экспорт завершён!\n\n"
        "  Квартир:    {}\n"
        "  Помещений:  {}\n\n"
        "Файл:\n{}".format(total_apt, total_rooms, save_path),
        title="Готово",
    )
    os.startfile(save_path)

except SystemExit:
    # sys.exit() — штатное завершение, не показываем ошибку
    pass

except KeyboardInterrupt:
    # Пользователь прервал выполнение
    pass

except Exception:
    import traceback
    alert(
        "Ошибка при экспорте:\n\n{}".format(traceback.format_exc()),
        title="Ошибка",
    )
